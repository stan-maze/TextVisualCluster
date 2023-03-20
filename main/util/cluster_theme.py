
from text_analysis_tools import KmeansClustering
from text_analysis_tools import DbscanClustering
from text_analysis_tools import CosionSimilarity
from text_analysis_tools import EditSimilarity
from text_analysis_tools import SimHashSimilarity
from text_analysis_tools import TfidfKeywords
from text_analysis_tools import TextRankKeywords
from text_analysis_tools import KeyPhraseExtraction
from text_analysis_tools import SentimentAnalysis
from text_analysis_tools import SpellCorrect
from text_analysis_tools import TfidfSummarization
from text_analysis_tools import TextRankSummarization
from text_analysis_tools import TopicKeywords
from text_analysis_tools import Fasttext
from text_analysis_tools import Word2VecSynonym
from text_analysis_tools import SynonymDict
from text_analysis_tools import TripleExtraction

import json
import os
import sys
import pandas as pd
import openpyxl

script_dir = os.path.dirname(os.path.abspath(__file__))

# Add a subdirectory of this directory to the module search path
subdir_path = os.path.join(script_dir, 'sentiment')
print(script_dir)
sys.path.append(subdir_path)


from util.sentiment.sentiment_analysis_dict.anatool import tool


def convert_to_excel_file(json_file_path):
    # 读取json数据
    with open(json_file_path, 'r', encoding='utf-8') as f:
        json_data = json.load(f)

    # 转换为DataFrame
    df = pd.DataFrame(json_data)
    df = df.explode('text')
    df = df.drop(columns=['num', 'mood'])
    df = df[['theme', 'text']]
    df = df.groupby(['theme'], as_index=False).agg(lambda x: '\n\n'.join(x))

    # 保存为Excel文件
    excel_file_path = os.path.splitext(json_file_path)[0] + '.xlsx'
    df.to_excel(excel_file_path, index=False, encoding='utf-8')
    print(f'Successfully converted {json_file_path} to {excel_file_path}')
    return excel_file_path



class ClusterTool:
    def __init__(self):
        pass


    def convert_to_excel_file(self, json_file_path):
        self.json_file_path = json_file_path
        self.xlsx_file_path = os.path.splitext(json_file_path)[0] + '.xlsx'
        with open(self.json_file_path, 'r', encoding='utf-8') as f:
            self.data = json.load(f)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Questions'
        sheet['A1'] = '关键词'
        sheet['B1'] = '问题'
        row_index = 2  # start from row 2
        current_theme = None
        for item in self.data:
            if item['theme'] != current_theme:
                current_theme = item['theme']
                row_index += 1  # skip one row for new theme
            for sentence in item['text']:
                sheet.cell(row=row_index, column=1, value=item['theme'])
                sheet.cell(row=row_index, column=2, value=sentence)
                row_index += 1
        workbook.save(self.xlsx_file_path)
        print(f'Successfully converted {self.json_file_path} to {self.xlsx_file_path}')
        return self.xlsx_file_path
    

    def textrank_summarization(self, doc, ratio=0.2):
        # with open(data_path, encoding="utf-8") as f:
        #     doc = f.read()
        summ = TextRankSummarization(ratio=ratio)
        summ = summ.analysis(doc)
        print("textrank summarization result: {}\n".format(summ))
        return summ


    def dbscan_cluster(self, data_path = '../resouce/test.xlsx',
                    eps=0.05, min_samples=3, fig=False):
        """
        基于DBSCAN进行文本聚类
        :param data_path: 文本路径，每行一条
        :param eps: DBSCA中半径参数
        :param min_samples: DBSCAN中半径eps内最小样本数目
        :param fig: 是否对降维后的样本进行画图显示, 默认False
        :return: {'cluster_0': [0, 1, 2, 3, 4], 'cluster_1': [5, 6, 7, 8, 9]}   0,1,2....为文本的行号
        """
        dbscan = DbscanClustering()
        result = dbscan.dbscan(corpus_path=data_path, eps=eps, min_samples=min_samples, fig=fig)
        return result
        
    def excuteCluster(self, txt_path,
                      eps=0.055, min_samples=3, fig=False):
        print(txt_path)
        result = self.dbscan_cluster(data_path = txt_path, eps=eps, min_samples=min_samples, fig=fig)
        
        with open(txt_path, "r", encoding="utf-8") as f:
            data = f.readlines()


        # 转换为 JSON 格式
        json_data = []
        mytool = tool()
        dic = {-1:'不好', 0:'中等', 1:'好'}
        for k, v in result.items():
            json_cluster = {}
            json_cluster["clusterId"] = int(k.split("_")[1])
            text = [data[i].strip() for i in v]
            topic_keywords = TopicKeywords(train_data=text, n_components=1,
                                    n_top_words=2, max_iter=15)
            s = '.'.join(text)
            
            # json_cluster["theme"] = self.textrank_summarization(s, ratio=0.2)
            
            json_cluster["theme"] = '|'.join(list(topic_keywords.analysis().values())[0])
            
            ppp = mytool.predict(s)
            print(ppp)
            json_cluster["mood"] = dic[ppp]
            json_cluster["num"] = len(v)
            json_cluster["text"] = text
            
            
            json_data.append(json_cluster)

        # 输出为 JSON 文件
        json_file_path = os.path.splitext(txt_path)[0] + '.json'
        with open(json_file_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f, ensure_ascii=False, indent=4)
            print(f'聚类成功，结果保存到 {json_file_path}')
        
        excel_file_path = self.convert_to_excel_file(json_file_path)
        return json_file_path, excel_file_path

    
   

def main():
    cluster_tool = ClusterTool()
    txt_path = '../resource/example.txt'
    eps = 0.1
    min_samples = 2
    fig = False
    result_file_path = cluster_tool.execute_cluster(txt_path, eps=eps, min_samples=min_samples, fig=fig)
    print(f"Results saved to {result_file_path}")


if __name__ == '__main__':
    main()
